from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import bcrypt
import os


def hash_password(plain: str) -> str:
    """Return a bcrypt hash string for the given plaintext password."""
    return bcrypt.hashpw(plain.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')


def verify_password(plain: str, hashed: str) -> bool:
    """Return True if plain matches the stored bcrypt hash."""
    try:
        return bcrypt.checkpw(plain.encode('utf-8'), hashed.encode('utf-8'))
    except Exception:
        return False

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app)

EXCEL_FILE = 'user.xlsx'

HEADERS = [
    "S.No", "First Name", "Middle Name", "Last Name",
    "Date of Birth", "Gender", "Languages", "Country",
    "Province/State", "District", "Municipality",
    "Mobile Number", "Email", "Photo", "Password (hashed)", "Registered At"
]

def style_header_row(ws):
    header_fill = PatternFill("solid", start_color="2D1B69")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin        = Side(style="thin", color="CCCCCC")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.row_dimensions[1].height = 36
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

def set_col_widths(ws):
    widths = [6, 14, 14, 14, 14, 10, 20, 14, 18, 14, 16, 16, 26, 14, 20, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        style_header_row(ws)
        set_col_widths(ws)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
        wb.save(EXCEL_FILE)
        print(f"Created {EXCEL_FILE}")
    else:
        wb = load_workbook(EXCEL_FILE)
        if "Users" not in wb.sheetnames:
            ws = wb.active
            ws.title = "Users"
            if ws.max_row == 1 and ws['A1'].value is None:
                style_header_row(ws)
                set_col_widths(ws)
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
            wb.save(EXCEL_FILE)

def get_next_sno(ws):
    return ws.max_row

def style_data_row(ws, row_num):
    fill   = PatternFill("solid", start_color="F3F0FF" if row_num % 2 == 0 else "FFFFFF")
    font   = Font(name="Arial", size=10, color="333333")
    center = Alignment(horizontal="center", vertical="center")
    thin   = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col in range(1, len(HEADERS) + 1):
        cell            = ws.cell(row=row_num, column=col)
        cell.font       = font
        cell.fill       = fill
        cell.alignment  = center
        cell.border     = border

@app.route('/')
def index():
    return app.send_static_file('index.html')

@app.route('/register', methods=['POST'])
def register():
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'message': 'No data received'}), 400

    required = ['firstName', 'lastName', 'dob', 'gender',
                'country', 'state', 'district', 'municipality', 'email', 'password']
    for field in required:
        if not data.get(field, '').strip():
            return jsonify({'success': False, 'message': f'{field} is required'}), 400

    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Users"]

        # Duplicate email check
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[12] and str(row[12]).lower() == data['email'].lower():
                return jsonify({'success': False, 'message': 'Email already registered'}), 409

        sno = get_next_sno(ws)
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        new_row = [
            sno,
            data.get('firstName', '').strip(),
            data.get('middleName', '').strip(),
            data.get('lastName', '').strip(),
            data.get('dob', '').strip(),
            data.get('gender', '').strip(),
            data.get('languages', '').strip(),
            data.get('country', '').strip(),
            data.get('state', '').strip(),
            data.get('district', '').strip(),
            data.get('municipality', '').strip(),
            data.get('phone', '').strip(),
            data.get('email', '').strip(),
            data.get('photo', '').strip(),
            hash_password(data.get('password', '').strip()),  # col 14 — bcrypt hash, never plaintext
            now                                               # col 15
        ]

        ws.append(new_row)
        style_data_row(ws, ws.max_row)
        wb.save(EXCEL_FILE)

        full_name = f"{data.get('firstName','')} {data.get('lastName','')}".strip()
        print(f"[{now}] Registered: {full_name} <{data.get('email','')}>")

        return jsonify({
            'success': True,
            'message': 'Registration successful!',
            'sno': sno,
            'name': full_name
        })

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/login', methods=['POST'])
def login():
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'message': 'No data received'}), 400

    email    = data.get('email', '').strip().lower()
    password = data.get('password', '').strip()

    if not email or not password:
        return jsonify({'success': False, 'message': 'Email and password are required'}), 400

    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
        ws = wb["Users"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            # row indices: [0]=SNo [1]=FirstName [2]=Mid [3]=Last [12]=Email [14]=PasswordHash
            stored_email = str(row[12]).strip().lower() if row[12] else ''
            stored_hash  = str(row[14]).strip()         if row[14] else ''

            if stored_email == email:
                if verify_password(password, stored_hash):
                    first_name = str(row[1]).strip() if row[1] else ''
                    last_name  = str(row[3]).strip() if row[3] else ''
                    full_name  = f"{first_name} {last_name}".strip()
                    return jsonify({
                        'success': True,
                        'message': f'Welcome back, {full_name}!',
                        'name': full_name,
                        'firstName': first_name,
                        'email': str(row[12])
                    })
                else:
                    return jsonify({'success': False, 'message': 'Incorrect password'}), 401

        return jsonify({'success': False, 'message': 'No account found with this email'}), 404

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/users', methods=['GET'])
def get_users():
    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
        ws = wb["Users"]
        users = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                users.append({
                    'sno':          row[0],
                    'firstName':    row[1],
                    'middleName':   row[2],
                    'lastName':     row[3],
                    'dob':          str(row[4]) if row[4] else '',
                    'gender':       row[5],
                    'languages':    row[6],
                    'country':      row[7],
                    'state':        row[8],
                    'district':     row[9],
                    'municipality': row[10],
                    'phone':        str(row[11]) if row[11] else '',
                    'email':        row[12],
                    'photo':        row[13],
                    'registeredAt': str(row[15]) if row[15] else '',
                })
        return jsonify({'success': True, 'users': users, 'count': len(users)})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


if __name__ == '__main__':
    init_excel()
    print("\n✅ Server running at http://localhost:5000")
    print(f"📄 Data saved to: {os.path.abspath(EXCEL_FILE)}\n")
    app.run(debug=True, port=5000)
